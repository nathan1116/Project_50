import sys
import time
import datetime
import pymssql
import pymysql
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.uic import loadUiType
import openpyxl
from PyQt5.uic.properties import QtCore, QtGui
from PyQt5 import QtCore, QtGui, QtWidgets
#from dbutil_20 import get_conn_20, close_conn_20
#from dbutil_50 import get_conn_50, close_conn_50

ui, _ = loadUiType('50_UT_JJ.ui')

class MainApp(QMainWindow, ui):
   
    # 定义构造
    def __init__(self):
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.handle_buttons()

    def handle_buttons(self):
        self.pihao.returnPressed.connect(self.data)
        self.Scan.returnPressed.connect(self.Barcode_Scan)
        self.chuansong_btn.clicked.connect(self.Barcode_add)
        self.hege_btn.clicked.connect(self.tijiao_hege)
        self.buhege_btn.clicked.connect(self.tijiao_buhege)
        self.daiding_btn.clicked.connect(self.tijiao_daiding)
        self.jilu_btn.clicked.connect(self.jilu)
        self.daiding_btn_R.clicked.connect(self.tijiao_pao)
        self.table_data.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_data.customContextMenuRequested.connect(self.youjian)

    def data(self):
        pihao = self.pihao.text()
        
        conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
        cur = conn50.cursor()
      
        try:
            sql = "select Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result,jianyanyuan,Time,Note from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' order by IndexNo"
            
            cur.execute(sql)
            table_data = cur.fetchall()
  
            row = len(table_data)
            vol = len(table_data[0])

            self.table_data.setRowCount(row)
            self.table_data.setColumnCount(vol)
            #QTableWidgetItem(str(table_data[8][6])).setBackground(QBrush(QColor("red")))
            for i in range(row):
                for j in range(vol):
                    data = QTableWidgetItem(str(table_data[i][j]))
                    data_1 = QTableWidgetItem(str(table_data[i][6])).text()

                    if  data_1 == '合格':
                        data.setBackground(QBrush(QColor("limegreen")))
                    if  data_1 == '不合格':
                        data.setBackground(QBrush(QColor("tomato")))
                    if  data_1 == '待定':
                        data.setBackground(QBrush(QColor("gold")))
                    
                    self.table_data.setItem(i,j,data)
                    data.setTextAlignment(Qt.AlignCenter)
                
            #self.table_data.resizeColumnsToContents()
            self.table_data.resizeRowsToContents()
            self.table_data.setAlternatingRowColors(True)
            self.table_data.scrollToBottom()
            conn50.close()
            #self.yueshu()
            self.tongji()
        except:
            QMessageBox.information(None, '错误', '数据预览出错,请检查服务器通讯！')
            conn50.close()

    def tongji(self):
        pihao = self.pihao.text()
        Time = time.strftime("%Y-%m-%d",time.localtime())
        conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
        cur = conn50.cursor()

        sql0 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测'"
        cur.execute(sql0)
        zs = cur.fetchall()
        for z in zs:
            z = list(z)
            z = z[0]
        self.LL.setText("来料："+str(z)+"支")

        sql1 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '合格'"
        cur.execute(sql1)
        hs = cur.fetchall()
        for h in hs:
            h = list(h)
            h = h[0]
        self.HG.setText("合格："+str(h)+"支")

        sql2 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '不合格'"
        cur.execute(sql2)
        bs = cur.fetchall()
        for b in bs:
            b = list(b)
            b = b[0]
        self.BHG.setText("不合格："+str(b)+"支")


        sql21 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '不合格' and InjuryDes like '%纵%'"
        cur.execute(sql21)
        bas = cur.fetchall()
        for ba in bas:
            ba = list(ba)
            ba = ba[0]
        self.BHG_ZS.setText("纵伤："+str(ba)+"支")

        sql22 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '不合格' and InjuryDes like '%草%' and InjuryDes like '%mm横%'"
        cur.execute(sql22)
        bbs = cur.fetchall()
        for bb in bbs:
            bb = list(bb)
            bb = bb[0]
        self.BHG_HS.setText("横伤："+str(bb)+"支")

        sql23 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '不合格' and InjuryDes like '%尺寸%'"
        cur.execute(sql23)
        bcs = cur.fetchall()
        for bc in bcs:
            bc = list(bc)
            bc = bc[0]
        self.BHG_CC.setText("尺寸超差："+str(bc)+"支")

        sql3 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '待定'"
        cur.execute(sql3)
        ds = cur.fetchall()
        for d in ds:
            d = list(d)
            d = d[0]
        self.DD.setText("待定："+str(d)+"支")


        sql0 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Time like '%"+str(Time)+"%'"
        cur.execute(sql0)
        zs = cur.fetchall()
        for z in zs:
            z = list(z)
            z = z[0]
        self.LL_2.setText("来料："+str(z)+"支")

        sql1 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '合格' and Time like '%"+str(Time)+"%'"
        cur.execute(sql1)
        hs = cur.fetchall()
        for h in hs:
            h = list(h)
            h = h[0]
        self.HG_2.setText("合格："+str(h)+"支")

        sql2 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '不合格' and Time like '%"+str(Time)+"%'"
        cur.execute(sql2)
        bs = cur.fetchall()
        for b in bs:
            b = list(b)
            b = b[0]
        self.BHG_2.setText("不合格："+str(b)+"支")


        sql21 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '不合格' and InjuryDes like '%纵%' and Time like '%"+str(Time)+"%'"
        cur.execute(sql21)
        bas = cur.fetchall()
        for ba in bas:
            ba = list(ba)
            ba = ba[0]
        self.BHG_2_ZS.setText("纵伤："+str(ba)+"支")

        sql22 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '不合格' and InjuryDes like '%草%' and  InjuryDes like '%mm横%' and Time like '%"+str(Time)+"%'"
        cur.execute(sql22)
        bbs = cur.fetchall()
        for bb in bbs:
            bb = list(bb)
            bb = bb[0]
        self.BHG_2_HS.setText("横伤："+str(bb)+"支")

        sql23 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '不合格' and InjuryDes like '%尺寸%' and Time like '%"+str(Time)+"%'"
        cur.execute(sql23)
        bcs = cur.fetchall()
        for bc in bcs:
            bc = list(bc)
            bc = bc[0]
        self.BHG_2_CC.setText("尺寸超差："+str(bc)+"支")

        sql3 = "select COUNT(*) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' and Result = '待定' and Time like '%"+str(Time)+"%'"
        cur.execute(sql3)
        ds = cur.fetchall()
        for d in ds:
            d = list(d)
            d = d[0]
        self.DD_2.setText("待定："+str(d)+"支")

        conn50.close()


    def tongji_Barcode(self):
     
        conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
        cur = conn50.cursor()

        sql0 = "select COUNT(*) from Tube"
        cur.execute(sql0)
        zs = cur.fetchall()
        for z in zs:
            z = list(z)
            z = z[0]
        self.shu.setText("当前管号数量："+str(z)+"支")

       
        conn50.close()


    def Barcode_Scan(self):
        global Barcode_1,Barcode_2
        pihao = self.pihao.text()
        
        ji = ['1', '3', '5', '7', '9']
        Barcode = self.Scan.text()
        conn20 = pymssql.connect(host="10.247.6.1",user="sa",password="xbg123!@#",database="XBG",charset="utf8")
        cur = conn20.cursor()

        sql = "select count(*) from Pipe where PipeCode = '"+str(Barcode)+"'"
        cur.execute(sql)
        nums = cur.fetchall()
        for num in nums:
            num = list(num)
            num = num[0]
        conn20.close()
        
        if num != 0:
            conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
            cur = conn50.cursor()

            sql = "select count(*) from Report where (Opername='JJ超声检测' and Barcode_1='"+str(Barcode)+"') or (Opername='JJ超声检测' and Barcode_2='"+str(Barcode)+"')"
            cur.execute(sql)
            lums = cur.fetchall()
            for lum in lums:
                lum = list(lum)
                lum = lum[0]
            conn50.close()

            if lum == 0:
                if list(Barcode)[9] in ji:
                    
                    Barcode_1 =  str(Barcode)
                    Barcode_2 = str(int(Barcode)+10)
                else:
                    Barcode_2 =  str(Barcode)
                    Barcode_1 = str(int(Barcode)-10)

                ###############################Barcode_all = str(Barcode_1)+' '+str(Barcode_2)

                ###########################self.show_Barcode.setText(Barcode_all)
                self.Barcode_wait()
                self.Scan.clear()
            else:
                QMessageBox.information(None, '错误', '管号重复,已经过超声检测！')
                self.Scan.clear()
        else:
            QMessageBox.information(None, '错误', '管号非法或不存在！')
            self.Scan.clear()



    def Barcode_wait(self):
   #     pihao = self.pihao.text()
        conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
        cur = conn50.cursor()
      
        try:
            sql0 = "select count(*) from tube where Barcode_1='"+str(Barcode_1)+"'"
            cur.execute(sql0)
            mas = cur.fetchall()
            for ma in mas:
                ma = list(ma)
                ma = ma[0]

            if ma == 0:
            
                sql1 = "insert into Tube (Barcode_1,Barcode_2) values('"+str(Barcode_1)+"','"+str(Barcode_2)+"')"
            
                sql = "select Barcode_1,Barcode_2 from Tube order by id"

            
                cur.execute(sql1)
                conn50.commit()

            
                cur.execute(sql)
                table_data = cur.fetchall()
  
                row = len(table_data)
                vol = len(table_data[0])

                self.table_Barcode.setRowCount(row)
                self.table_Barcode.setColumnCount(vol)
                #QTableWidgetItem(str(table_data[8][6])).setBackground(QBrush(QColor("red")))
                for i in range(row):
                    for j in range(vol):
                        data = QTableWidgetItem(str(table_data[i][j]))
                        
                        self.table_Barcode.setItem(i,j,data)
                        data.setTextAlignment(Qt.AlignCenter)
                    
                #self.table_data.resizeColumnsToContents()
                self.table_Barcode.resizeRowsToContents()
                self.table_Barcode.setAlternatingRowColors(True)
                self.table_Barcode.scrollToBottom()
                conn50.close()
                #self.yueshu()
                self.tongji_Barcode()               
        except:
            QMessageBox.information(None, '错误', '扫码出错,请检查服务器通讯！')
            conn50.close()


    def Barcode_add(self):
        global current_1,current_2


        conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
        cur = conn50.cursor()
  
        try:
            cur.execute("select Barcode_1,Barcode_2 from Tube order by id limit 1")
            current = cur.fetchall()
            current_1 = current[0][0]
            current_2 = current[0][1]
            
            
            Barcode_all = str(current_1)+' '+str(current_2)

            self.show_Barcode.setText(Barcode_all)

            cur.execute("delete from Tube where Barcode_1='"+str(current_1)+"'")
            conn50.commit()
         
        

            cur.execute("select Barcode_1,Barcode_2 from Tube order by id")
      
            table_data = cur.fetchall()
      
            row = len(table_data)
            vol = len(table_data[0])

            self.table_Barcode.setRowCount(row)
            self.table_Barcode.setColumnCount(vol)
            
            for i in range(row):
                for j in range(vol):
                    data = QTableWidgetItem(str(table_data[i][j]))
                        
                    self.table_Barcode.setItem(i,j,data)
                    data.setTextAlignment(Qt.AlignCenter)

            self.table_Barcode.resizeRowsToContents()
            self.table_Barcode.setAlternatingRowColors(True)
            self.table_Barcode.scrollToBottom()
            conn50.close()
    
            self.tongji_Barcode()

        except:
            QMessageBox.information(None, '错误', '管号已用完！')
            self.table_Barcode.clear()
            

            self.tongji_Barcode()
            
            conn50.close()


    def tijiao_hege(self):
        pihao = self.pihao.text()
        jianyanyuan = self.jianyanyuan.text()
        banci = self.banci.currentText()
        biaozhun = self.biaozhun.text()
        local_time = time.strftime("%Y-%m-%d",time.localtime())
        Time = local_time+" "+str(banci)
        a = self.show_Barcode.text()
     
        if a != '':
            try:
                conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
                cur = conn50.cursor()
            
                cur.execute("select max(IndexNo) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测'")
           
                maxs = cur.fetchall()
                print(maxs)
                for max_ in maxs:
                    max_ = list(max_)
                    max_ = max_[0]

                if max_ is not None:
                    IndexNo = str(int(max_)+1).zfill(3)
                else:
                    IndexNo = "001"

                sql = "insert into Report (Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result,jianyanyuan,Time,Note) values ('JJ超声检测','"+str(pihao)+"','"+str(IndexNo)+"','"+str(current_1)+"','"+str(current_2)+"','无超标缺陷 尺寸合格','合格','"+str(jianyanyuan)+"','"+str(Time)+"','"+str(biaozhun)+"')"
               
                cur.execute(sql)
                conn50.commit()
                
                self.data()
                self.show_Barcode.clear()
                self.Barcode_add()
                self.biaozhun.clear()
                conn50.close()
            except:
                QMessageBox.information(None, '错误', '管号或其他信息有误！')
                self.show_Barcode.clear()
                self.biaozhun.clear()
                conn50.close()

        else:
            QMessageBox.information(None, '错误', '未发现待检管号！')
            

    def tijiao_buhege(self):
        pihao = self.pihao.text()
        jianyanyuan = self.jianyanyuan.text()
        banci = self.banci.currentText()
        biaozhun = self.biaozhun.text()
        shangqing = self.shangqing.text()
        local_time = time.strftime("%Y-%m-%d",time.localtime())
        Time = local_time+" "+str(banci)
        a = self.show_Barcode.text()
        if a != '':
            try:
                conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
                cur = conn50.cursor()
            
                cur.execute("select max(IndexNo) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测'")
                maxs = cur.fetchall()

                for max_ in maxs:
                    max_ = list(max_)
                    max_ = max_[0]

                if max_ is not None:
                    IndexNo = str(int(max_)+1).zfill(3)
                else:
                    IndexNo = "001"

                sql = "insert into Report (Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result,jianyanyuan,Time,Note) values ('JJ超声检测','"+str(pihao)+"','"+str(IndexNo)+"','"+str(current_1)+"','"+str(current_2)+"','"+str(shangqing)+"','不合格','"+str(jianyanyuan)+"','"+str(Time)+"','"+str(biaozhun)+"')"
                           
                cur.execute(sql)
                conn50.commit()
                self.data()
                self.show_Barcode.clear()
                self.Barcode_add()
                self.biaozhun.clear()
                self.shangqing.clear()
                conn50.close()
            except:
                QMessageBox.information(None, '错误', '管号或其他信息有误！')
                self.show_Barcode.clear()
                self.biaozhun.clear()
                self.shangqing.clear()
                conn50.close()

        else:
            QMessageBox.information(None, '错误', '未发现待检管号！')

    def tijiao_daiding(self):
        pihao = self.pihao.text()
        jianyanyuan = self.jianyanyuan.text()
        banci = self.banci.currentText()
        biaozhun = self.biaozhun.text()
        shangqing = self.shangqing.text()
        local_time = time.strftime("%Y-%m-%d",time.localtime())
        Time = local_time+" "+str(banci)
        a = self.show_Barcode.text()
        if a != '':
            try:
                conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
                cur = conn50.cursor()
            
                cur.execute("select max(IndexNo) from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测'")
                maxs = cur.fetchall()

                for max_ in maxs:
                    max_ = list(max_)
                    max_ = max_[0]

                if max_ is not None:
                    IndexNo = str(int(max_)+1).zfill(3)
                else:
                    IndexNo = "001"

                sql = "insert into Report (Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result,jianyanyuan,Time,Note) values ('JJ超声检测','"+str(pihao)+"','"+str(IndexNo)+"','"+str(current_1)+"','"+str(current_2)+"','尺寸或草状显示','待定','"+str(jianyanyuan)+"','"+str(Time)+"','"+str(biaozhun)+"')"
                           
                cur.execute(sql)
                conn50.commit()
                self.data()
                self.show_Barcode.clear()
                self.Barcode_add()
                self.biaozhun.clear()
                conn50.close()
            except:
                QMessageBox.information(None, '错误', '管号或其他信息有误！')
                self.show_Barcode.clear()
                self.biaozhun.clear()
                conn50.close()
           
        else:
            QMessageBox.information(None, '错误', '未发现待检管号！')

    def pao(self):
        global Pipe_all
        Pipe_all = []
        pihao = self.pihao.text()

        conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
        cur = conn50.cursor()

        sql = "select Barcode_1,Barcode_2 from Report where pihao = '" + str(pihao) + "' and Opername like '%超声%' and Result = '待定'"

        cur.execute(sql)
        Pipe = cur.fetchall()
       
        for i in list(Pipe):
            Pipe_all.append(i[0])
            Pipe_all.append(i[1])
        conn50.close()

    def tijiao_pao(self):
        self.pao()
    
        pihao = self.pihao.text()
        conn20 = pymssql.connect(host="10.247.6.1",user="sa",password="xbg123!@#",database="XBG",charset="utf8")
        cur = conn20.cursor()

        #重复提交？
      
        try:
            for i in Pipe_all:
                sql = "delete from ReLaserMark where PipeCode='"+str(i)+"'"
                sql1 = "insert into Rework (BN,PipeCode,reason_5,reason_9,rework_2,rework_3,Operator_AID,CurPFID,RWState,SN,RWDID,FileIndex) values ('"+str(pihao)+"','"+str(i)+"','1','0','1','1','28','2','0','XXX','XXX','')"

                cur.execute(sql)
                conn20.commit()
                
                cur.execute(sql1)
                conn20.commit()
            QMessageBox.information(None, '成功', '管号成功导入抛光工序！')
            conn20.close()
     
        except:
            QMessageBox.information(None, '错误', '管号导入错误！')
            conn20.close()
        
    def youjian(self,pos):

        for i in self.table_data.selectionModel().selection().indexes():
            row_num = i.row()

        menu = QMenu()
        item1 = menu.addAction("删除本条记录")
        item2 = menu.addAction("改为合格")
        item3 = menu.addAction("改为不合格")
        item4 = menu.addAction("改为待定返工")
        action = menu.exec_(self.table_data.mapToGlobal(pos))

        if action == item1:
            pihao = self.pihao.text()
            conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
            cur = conn50.cursor()
            sql = "select Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result,jianyanyuan,Time,Note from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' order by IndexNo"
            cur.execute(sql)
            table_data = cur.fetchall()
            dele = table_data[row_num][2]
            dele = str(int(dele)).zfill(3)
            sql = "delete from Report where pihao = '"+str(pihao)+"' and IndexNo = '"+str(dele)+"' and Opername= 'JJ超声检测' "
            cur.execute(sql)
            conn50.commit()
            conn50.close()
            self.data()

        if action == item2:
            pihao = self.pihao.text()  
            conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
            cur = conn50.cursor()
            sql = "select Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result,jianyanyuan,Time,Note from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' order by IndexNo"         
            cur.execute(sql)
            table_data = cur.fetchall()
            upd = table_data[row_num][2]         
            upd = str(int(upd)).zfill(3)              
            sql = "update Report set  Result = '合格',InjuryDes='无超标缺陷 尺寸合格' where pihao = '"+str(pihao)+"' and IndexNo = '"+str(upd)+"' and Opername= 'JJ超声检测'"
            cur.execute(sql)
            conn50.commit()
            conn50.close()
            self.data()

        if action == item3:
            pihao = self.pihao.text()
            shangqing = self.shangqing.text()
            conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
            cur = conn50.cursor()
            sql = "select Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result,jianyanyuan,Time,Note from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' order by IndexNo"        
            cur.execute(sql)
            table_data = cur.fetchall()
            upd = table_data[row_num][2]            
            upd = str(int(upd)).zfill(3)           
            sql = "update Report set  Result = '不合格',InjuryDes='"+str(shangqing)+"' where pihao = '"+str(pihao)+"' and IndexNo = '"+str(upd)+"' and Opername= 'JJ超声检测'"
            cur.execute(sql)
            conn50.commit()
            conn50.close()
            self.data()

        if action == item4:
            pihao = self.pihao.text()
            conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
            cur = conn50.cursor()
            sql = "select Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result,jianyanyuan,Time,Note from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' order by IndexNo"
            cur.execute(sql)
            table_data = cur.fetchall()
            upd = table_data[row_num][2]
            upd = str(int(upd)).zfill(3)
            sql = "update Report set  Result = '待定',InjuryDes='尺寸或草状显示' where pihao = '"+str(pihao)+"' and IndexNo = '"+str(upd)+"' and Opername= 'JJ超声检测'"
            cur.execute(sql)
            conn50.commit()
            conn50.close()
            self.data()


    def jilu(self):
        pihao = self.pihao.text()
        wb = openpyxl.load_workbook("UTMB.xlsx")
        ws = wb['Sheet1']

        m = 13
        p = 0
        n = 35

        conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
        cur = conn50.cursor()
      
        sql = "select IndexNo,InjuryDes,Result,jianyanyuan,Time,Barcode_1,Barcode_2  from Report where pihao = '"+str(pihao)+"' and Opername= 'JJ超声检测' order by IndexNo"
        
        cur.execute(sql)
        datas = cur.fetchall()
  
        for data in datas:

            if m <= 30:
        
                ws.cell(row=m, column=1).value = data[0]
                ws.cell(row=m, column=2).value = data[1]
                ws.cell(row=m, column=4).value = data[2]
                ws.cell(row=m, column=5).value = data[3]
                ws.cell(row=m, column=6).value = data[4]
                ws.cell(row=m, column=7).value = ''.join(list(data[5])[6:11])+'/'+''.join(list(data[6])[6:11])

                m = m + 1
                
            else:
                if p == 25:
                
                    p = 0
                    n = n + 5 
             
                ws.cell(row=n, column=1).value = data[0]
                ws.cell(row=n, column=2).value = data[1]
                ws.cell(row=n, column=4).value = data[2]
                ws.cell(row=n, column=5).value = data[3]
                ws.cell(row=n, column=6).value = data[4]
                ws.cell(row=n, column=7).value = ''.join(list(data[5])[6:11])+'/'+''.join(list(data[6])[6:11])

                n = n + 1
                p = p + 1
                
        wb.save("C:/Users/Administrator/Desktop/检验记录/JJ超声记录"+str(pihao)+".xlsx")

        conn50.close()

        wb = openpyxl.load_workbook("C:/Users/Administrator/Desktop/检验记录/JJ超声记录"+str(pihao)+".xlsx")
        ws = wb['Sheet1']
        for i in range(100):
            n = 30*i + 35
            m = ws.cell(row=n, column=1).value
            if m is None:
                break;
        for i in range(100):
            ws.delete_rows(n-4,30)

        wb.save("C:/Users/Administrator/Desktop/检验记录/JJ超声记录"+str(pihao)+".xlsx")
        
        QMessageBox.information(None,'完成','已生成记录，存放在桌面文件夹内！')

       

def main():
    app = QApplication(sys.argv)
    splash = QSplashScreen(QPixmap("1.png"))
    splash.setFont(QFont('Microsoft YaHei UI', 15))
    splash.show()
    splash.showMessage("启动中...",QtCore.Qt.AlignHCenter |QtCore.Qt.AlignBottom, QtCore.Qt.white)
    time.sleep(0.2)
    ###splash.showMessage("正在连接数据库...",QtCore.Qt.AlignHCenter |QtCore.Qt.AlignBottom, QtCore.Qt.white)
    ###time.sleep(0.2)
    ###conn50 = get_conn_50()
    ###if (conn50 != 0):
     ###   conn50.close()
      ###  splash.showMessage("数据库连接成功...",QtCore.Qt.AlignHCenter |QtCore.Qt.AlignBottom, QtCore.Qt.white)
   ### else:
      ###  splash.showMessage("数据库连接失败...",QtCore.Qt.AlignHCenter |QtCore.Qt.AlignBottom, QtCore.Qt.white)
    # time.sleep(0.5)
    #app = QApplication([])
    window = MainApp()
    window.show()
    splash.finish(window)
    app.exec_()

if __name__ == '__main__':
    main()

