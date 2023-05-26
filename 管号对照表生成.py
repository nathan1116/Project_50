import pymysql
from tkinter import filedialog
from tkinter import *
from tkinter import ttk
import tkinter as tk
import tkinter.font as tkFont
import openpyxl
import time
from openpyxl.styles import Font
from openpyxl.styles import PatternFill,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


root=Tk()
root.title("管号对照表")
root.geometry('410x100')

var1=tk.StringVar()
var2=tk.StringVar()
var3=tk.StringVar()
var4=tk.StringVar()

ft1=tkFont.Font(size=15,weight=tkFont.BOLD,slant=tkFont.ROMAN)
ft2=tkFont.Font(size=15,weight=tkFont.BOLD,slant=tkFont.ROMAN)
ft3=tkFont.Font(size=15,weight=tkFont.BOLD,slant=tkFont.ROMAN)
Label(root,text=str("*生产批号 "),padx=5,font =ft1).grid(row=0,column=0,padx=1,pady=10)


e1=Entry(root,borderwidth=1,width=10,foreground='red',font=('Helvetica','15','bold'))
e1.grid(row=0,column=1,padx=1,pady=10)


def chaxun():
    Filepath=filedialog.askdirectory()
    #Filepath=filedialog.askopenfilename()
    pihao=e1.get()
    pihao="'"+str(pihao)+"'"
    #print(pihao)

    pipi=e1.get()
    wb=openpyxl.Workbook()
    ws=wb['Sheet']
    ws.title='管号清单'
    TM=time.strftime("%Y-%m-%d %H-%M",time.localtime())   

    conn50 = pymysql.connect(host='192.168.100.2',database='50',user='root',password='123456')
    cur = conn50.cursor()

    #sqlvt="SELECT Barcode FROM MTSPro_ReportingInfo where Opername='外观检验' and ProBatchNo="\
       #  +str(pihao)+" and Status='1' ORDER BY IndexNo"

    sqlvt="select Opername,pihao,IndexNo,Barcode_1,Barcode_2,InjuryDes,Result  from Report where pihao = "+str(pihao)+" order by IndexNo"
 
   
            
    cur.execute(sqlvt)
    results=cur.fetchall()
  
    for result in results:
        ws.append(result)        


 
    bian=Border(left=Side(border_style='thin',color='000000'),
          right=Side(border_style='thin',color='000000'),
          top=Side(border_style='thin',color='000000'),
          bottom=Side(border_style='thin',color='000000'))
    

      
    Filepathv=Filepath+'/'+str(pihao)+' '+str(TM)+'.xlsx'
    print (Filepathv)
    
    wb.save(Filepathv)
  
    
    conn50.close()
    print("已生成！")


Button(root,text="生成Excel",font =ft1,width=9,command=chaxun).grid(row=0,column=3,padx=10,pady=5)

mainloop()
